from flask import Flask, request, send_file, jsonify
import openpyxl
import re
import os

app = Flask(__name__)

# Función para extraer datos usando expresiones regulares
def extract_data_with_regex(text, pattern):
    match = re.search(pattern, text, re.DOTALL)
    return match.group(1).strip() if match else 'No encontrado'

# Función para limpiar texto específico de un campo
def clean_text(text, field_name):
    cleaned_text = text.strip()
    if field_name in ['Cargo', 'Asunto', 'Entidad']:
        cleaned_text = cleaned_text.split('\n')[0].strip()
    return cleaned_text

# Función para actualizar un archivo Excel con datos extraídos
def update_excel_with_data(file_path, extracted_data):
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames
        print(f"Hoja encontrada: {sheet_names}")

        if not sheet_names:
            raise ValueError('El archivo Excel no contiene hojas.')

        ws = wb[sheet_names[0]]

        headers = {
            'Fecha de Radicación': 2,
            'Hora de Radicación': 3,
            'Fecha de Petición': 4,
            'Hora de Petición': 5,
            '# de Petición': 6,
            'Solicitante': 7,
            'Correo de Solicitante': 8,
            'Entidad': 9,
            'Cargo': 10,
            'Asunto': 11
        }

        # Encuentra la primera fila vacía
        row = 2
        while any(ws.cell(row=row, column=col).value for col in headers.values()):
            row += 1

        for key, col in headers.items():
            ws.cell(row=row, column=col, value=extracted_data.get(key, 'No encontrado'))

        updated_file_path = file_path.replace('.xlsx', '_updated.xlsx')
        wb.save(updated_file_path)

        return updated_file_path

    except Exception as e:
        print(f'Error al actualizar el archivo Excel: {e}')
        raise

@app.route('/update', methods=['POST'])
def update_file():
    if 'file' not in request.files or 'text' not in request.form:
        return 'Texto o archivo Excel no proporcionado.', 400

    file = request.files['file']
    text = request.form['text']
    if file.filename == '':
        return 'No seleccionó ningún archivo.', 400

    # Guardar en la carpeta temporal
    temp_file_path = '/tmp/' + file.filename
    file.save(temp_file_path)

    # Procesar y actualizar el archivo
    patterns = {
        'Fecha de Radicación': r'Fecha de Radicación:\s*(\d{2}/\d{2}/\d{4})',
        'Hora de Radicación': r'Fecha de Radicación:\s*\d{2}/\d{2}/\d{4} (\d{1,2}:\d{2} [ap]m)',
        'Fecha de Petición': r'Date:\s*\w{3}, (\d{2} \w{3} \d{4})',
        'Hora de Petición': r'Date:\s*\w{3}, \d{2} \w{3} \d{4} a las (\d{1,2}:\d{2})',
        '# de Petición': r'radicado/consecutivo ([^\s]+)',
        'Solicitante': r'Nombres:\s*([\w\s]+)',
        'Correo de Solicitante': r'Email:\s*([\w\.-]+@[\w\.-]+)',
        'Entidad': r'Empresa:\s*([^\n]+)',
        'Cargo': r'Cargo:\s*([^\n]+)',
        'Asunto': r'Asunto:\s*([^\n]+)'
    }

    extracted_data = {}
    for key, pattern in patterns.items():
        extracted_data[key] = extract_data_with_regex(text, pattern)

    updated_file_path = update_excel_with_data(temp_file_path, extracted_data)

    # Guardar la ruta del archivo actualizado en algún lugar (en la memoria, en una base de datos, etc.)
    # Aquí guardamos el nombre del archivo en un archivo simple para referencia
    with open('/tmp/last_update.txt', 'w') as f:
        f.write(updated_file_path)

    # Enviar respuesta de éxito
    return jsonify({'message': 'Archivo actualizado con éxito', 'file_path': updated_file_path})

@app.route('/download', methods=['GET'])
def download_file():
    try:
        with open('/tmp/last_update.txt', 'r') as f:
            updated_file_path = f.read().strip()
        if not os.path.exists(updated_file_path):
            return 'Archivo no encontrado.', 404
        return send_file(updated_file_path, as_attachment=True)
    except Exception as e:
        return str(e), 500

@app.route('/')
def index():
    return app.send_static_file('index.html')

if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(host='0.0.0.0', port=8080)
